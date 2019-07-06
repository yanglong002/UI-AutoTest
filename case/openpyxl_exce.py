# openpyxl  第三方库  可对excel进行读写

# 安装方法： pip install openpyxl

#  只支持xlsx格式

from openpyxl import load_workbook #读写
from openpyxl import Workbook #创建

# 1.创建一个excel
# wb = Workbook()
# wb.save('python.xlsx')

# 2.对excel读写
wb = load_workbook('python.xlsx')
sheet = wb['test']
max_row = sheet.max_row
max_column = sheet.max_column
print(max_column)
print(max_row)

for i in range(1, max_row+1):
    for j in range(1, max_column):
        res = sheet.cell(i, j).value
        print('{}行{}列的值是:{},类型为:{}'.format(i, j, res, type(res)))



