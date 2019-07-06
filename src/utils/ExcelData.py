from openpyxl import load_workbook
from openpyxl import Workbook

class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.a = None
        self.b = None
        self.expected = None


class DoExcelData:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    # 方法1
    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2, sheet.max_row+1):
            tmp = {}
            tmp['case_id'] = sheet.cell(i, 1).value
            tmp['title'] = sheet.cell(i, 2).value
            tmp['a']= sheet.cell(i, 3).value
            tmp['b'] = sheet.cell(i, 4).value
            tmp['expected'] = sheet.cell(i, 5).value
            test_data.append(tmp)
        return test_data
    # 方法2 类和对象的思想 需熟悉这种写法
    def get_data2(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        cases = []
        for i in range(2, sheet.max_row+1):
            case = Case()
            case.case_id = sheet.cell(i, 1).value
            case.title = sheet.cell(i, 2).value
            case.a= sheet.cell(i, 3).value
            case.b = sheet.cell(i, 4).value
            case.expected = sheet.cell(i, 5).value
            cases.append(case)
        return cases

    def write_back(self, row, col, value):
        wb = load_workbook(self.file_name) # 打开excel
        sheet = wb[self.sheet_name] # 指定页
        sheet.cell(row, col).value = value # 在指定行、列写入数据
        wb.save(self.file_name) # 写入后要保存,此时要关闭excel



if __name__ == "__main__":
    from config import configInfo as cf
    import os
    # 方法1
    file_name =os.path.join(cf.data_path, 'python.xlsx')
    test_data = DoExcelData(file_name, 'test').get_data()
    print(test_data)

    # 方法2 用类和对象的思想
    file_name = os.path.join(cf.data_path, 'python.xlsx')
    cases = DoExcelData(file_name, 'test').get_data2()
    for case in cases:
        print(case.case_id, case.title, case.a, case.b, case.expected)
